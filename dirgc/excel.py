import math
import os

from .settings import (
    DEFAULT_EXCEL_FILE,
    LEGACY_EXCEL_FILE,
    VALID_HASIL_GC_CODES,
)


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    return str(value).strip()


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""
    return " ".join(text.split())


def header_matches(header, name):
    if not header:
        return False
    name = name.lower()
    if header == name:
        return True
    if header.startswith(name + " "):
        return True
    if header.startswith(name + ":"):
        return True
    return False


def normalize_lat_lon(value, min_value, max_value):
    if value is None:
        return ""
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    if math.isnan(number):
        return ""
    if number < min_value or number > max_value:
        return ""
    if number.is_integer():
        return str(int(number))
    return str(number)


def normalize_code(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def normalize_hasil_gc(value):
    code = normalize_code(value)
    if code in VALID_HASIL_GC_CODES:
        return code
    return None


def resolve_excel_path(excel_file):
    if excel_file:
        return os.path.expanduser(excel_file)

    candidates = [
        os.path.join(os.getcwd(), DEFAULT_EXCEL_FILE),
        os.path.join(os.getcwd(), LEGACY_EXCEL_FILE),
    ]
    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate

    raise FileNotFoundError(
        "Excel file not found. Use --excel-file or place "
        f"{DEFAULT_EXCEL_FILE} in the current directory."
    )


def load_excel_rows(excel_path):
    path = resolve_excel_path(excel_path)

    try:
        import pandas as pd
    except ImportError:
        pd = None

    if pd:
        try:
            df = pd.read_excel(path, dtype=str)
        except Exception:
            df = None
        if df is None:
            pd = None
    if pd:
        columns = [normalize_header(col) for col in df.columns]

        def find_col(*names):
            for name in names:
                for index, column in enumerate(columns):
                    if header_matches(column, name):
                        return df.columns[index]
            return None

        col_idsbr = find_col("idsbr")
        col_nama = find_col("nama_usaha", "nama usaha", "namausaha", "nama")
        col_alamat = find_col("alamat", "alamat usaha", "alamat_usaha")
        col_lat = find_col("latitude", "lat")
        col_lon = find_col("longitude", "long", "lon")
        col_hasil = find_col(
            "hasil_gc",
            "hasil gc",
            "hasilgc",
            "ag",
            "keberadaanusaha_gc",
        )

        if col_hasil is None and df.shape[1] >= 33:
            col_hasil = df.columns[32]

        rows = []
        for _, row in df.iterrows():
            record = {
                "idsbr": normalize_text(row[col_idsbr]) if col_idsbr else "",
                "nama_usaha": normalize_text(row[col_nama]) if col_nama else "",
                "alamat": normalize_text(row[col_alamat]) if col_alamat else "",
                "latitude": normalize_lat_lon(row[col_lat], -90, 90)
                if col_lat
                else "",
                "longitude": normalize_lat_lon(row[col_lon], -180, 180)
                if col_lon
                else "",
                "hasil_gc": normalize_hasil_gc(row[col_hasil])
                if col_hasil is not None
                else None,
            }
            if any([record["idsbr"], record["nama_usaha"], record["alamat"]]):
                rows.append(record)
        return rows

    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Install pandas or openpyxl to read Excel files."
        ) from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = [normalize_header(cell.value) for cell in sheet[1]]

        def header_index(*names):
            for name in names:
                for index, header in enumerate(headers):
                    if header_matches(header, name):
                        return index + 1
            return None

        col_idsbr = header_index("idsbr")
        col_nama = header_index("nama_usaha", "nama usaha", "namausaha", "nama")
        col_alamat = header_index("alamat", "alamat usaha", "alamat_usaha")
        col_lat = header_index("latitude", "lat")
        col_lon = header_index("longitude", "long", "lon")
        col_hasil = header_index(
            "hasil_gc",
            "hasil gc",
            "hasilgc",
            "ag",
            "keberadaanusaha_gc",
        )
        if col_hasil is None and sheet.max_column >= 33:
            col_hasil = 33

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            def cell_value(col_index):
                if not col_index:
                    return None
                if col_index - 1 >= len(row):
                    return None
                return row[col_index - 1]

            record = {
                "idsbr": normalize_text(cell_value(col_idsbr)),
                "nama_usaha": normalize_text(cell_value(col_nama)),
                "alamat": normalize_text(cell_value(col_alamat)),
                "latitude": normalize_lat_lon(cell_value(col_lat), -90, 90),
                "longitude": normalize_lat_lon(cell_value(col_lon), -180, 180),
                "hasil_gc": normalize_hasil_gc(cell_value(col_hasil))
                if col_hasil
                else None,
            }
            if any([record["idsbr"], record["nama_usaha"], record["alamat"]]):
                rows.append(record)
        return rows
    finally:
        workbook.close()
